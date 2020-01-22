from project_gui.inventory_gui import *

from player_inventory import *

from project_gui.battle_gui import *

from project_gui.windows_for_many_cases import *

from ingame_data.json_data_upload import *

from ingame_data.player_data.json_player_data_upload import *

from parameters import *

from random import choice

import parameters as parameters_script

import tkinter as tk

import openpyxl

import os


layered = pg.sprite.LayeredUpdates()         # Доп группа для выбора приоритетных спрайтов
player_group = pg.sprite.GroupSingle()       # Игрок
all_sprites_group = pg.sprite.Group()        # Все спрайты
wall_like_sprites = pg.sprite.Group()        # Стены
gate_like_sprites = pg.sprite.Group()        # Врата
sign_like_sprites = pg.sprite.Group()        # Таблички
movable_objects = pg.sprite.Group()          # Коробки
save_points = pg.sprite.Group()              # Точки сохранения
chests = pg.sprite.Group()                   # Сундуки
enemies = pg.sprite.Group()                  # Враги
change_position_things = pg.sprite.Group()   # Объекты смены позиции игрока


def load_and_play(song, param=None):

    if pg.mixer.get_busy():

        pg.mixer.stop()

    if param:

        pg.mixer.music.set_volume(param * parameters_thing.VOLUME)

    pg.mixer.music.load(song)

    pg.mixer.music.play(-1)


class Player(pg.sprite.Sprite):
    """
    Потомок класса pg.sprite.Sprite.
    Здесь описано всё, что связано с созданием спрайта типа "Персонаж".
    Полностью описаны все типа взаимодействия с объектами окружающей среды.
    ( Не совсем опрятно, но не суть )
    Показатели и остальные данные берутся с JSON файла при использовании методов класса PlayerDataJSON.
    Для удобства хранения данных берутся экземпляры классов Inventory и NumericStats.
    """
    def __init__(self, x, y, group=layered):

        pg.sprite.Sprite.__init__(self, group)

        self._layer = 0

        self.image_animation = [pg.image.load(os.path.join(os.getcwd(), 'ingame_data',
                                                           'player_data', 'player_textures', 'player_texture_0.png')),
                                pg.image.load(os.path.join(os.getcwd(), 'ingame_data',
                                                           'player_data', 'player_textures', 'player_texture_1.png')),
                                pg.image.load(os.path.join(os.getcwd(), 'ingame_data',
                                                           'player_data', 'player_textures', 'player_texture_2.png'))
                                ]

        self.image = self.image_animation[0]

        self.idx = 0

        self.rect = self.image.get_rect()

        self.collide_rect = pg.Rect(0, 0, TILESIZE, TILESIZE)

        self.dx, self.dy = 0, 0

        self._inventory = Inventory()  # FIXME: и это

        self._numeric_stats = NumericStats(100, 0)

        self._log_chests = []

        self.collide_rect.topleft = (x, y)

    def update(self):

        self.rect.bottomleft = self.collide_rect.bottomleft

        self._operate_with_keys()

        layered.move_to_front(self)

    def _movement(self):

        if self.idx < len(self.image_animation):

            self.image = self.image_animation[self.idx]

            self.idx += 1

        else:

            self.idx = 0

        self.collide_rect.x += self.dx

        self.collide_rect.y += self.dy

        self._detect_collision()

    def _operate_with_keys(self):

        self.dx, self.dy = 0, 0

        self.image = self.image_animation[0]

        keys = pg.key.get_pressed()

        if keys[parameters_thing.TAKE_PHOTO]:

            parameters_thing.SCREENSHOT = True

        if keys[parameters_thing.OPEN_INV]:

            top = tk.Tk()

            InventoryGUI(top, {'NUMERIC': self._numeric_stats.get_numeric,
                               'INVENTORY': self._inventory.get_items}).loop()

        if keys[parameters_thing.LEFT]:

            self.dx, self.dy = -VELOCITY, 0

            self._movement()

        elif keys[parameters_thing.RIGHT]:

            self.dx, self.dy = VELOCITY, 0

            self._movement()

        if keys[parameters_thing.UP]:

            self.dx, self.dy = 0, -VELOCITY

            self._movement()

        elif keys[parameters_thing.DOWN]:

            self.dx, self.dy = 0, VELOCITY

            self._movement()

    def _collision(self, sprite):

        if self.collide_rect.colliderect(sprite.collide_rect):

            if self.dx > 0:
                self.collide_rect.right = sprite.collide_rect.left

            if self.dx < 0:
                self.collide_rect.left = sprite.collide_rect.right

            if self.dy > 0:
                self.collide_rect.bottom = sprite.collide_rect.top

            if self.dy < 0:
                self.collide_rect.top = sprite.collide_rect.bottom

    def _enemy_collision(self, sprite):

        if self.collide_rect.colliderect(sprite.collide_rect):

            top = Tk()

            load_and_play(os.path.join(os.getcwd(), 'soundtracks', 'battle.mp3'), 0.6)

            if BattleGUI(top, self._numeric_stats.get_numeric, self._inventory.get_items, sprite.enemy_type).loop():

                sprite.kill()

            else:

                parameters_thing.DEATH = True

            load_and_play(parameters_script.CURRENT_SONG)

            self._collision(sprite)

    def _chest_collision(self, sprite):

        if self.collide_rect.colliderect(sprite.collide_rect):

            top = Tk()

            if sprite._item is not None:

                if self._inventory:

                    InformationWindow(top, 'Найден предмет {}'.format(sprite._item['NAME']),
                                      'Информация', 'Закрыть')

                    self._inventory.append(sprite._item)

                    sprite.empty()

                    self._log_chests.append(sprite.collide_rect.topleft)

                else:

                    InformationWindow(top, 'Ваш инвентарь полностью заполнен!',
                                      'Информация', 'Закрыть')

            else:

                InformationWindow(top, 'Сундук пуст',
                                  'Информация', 'Закрыть')

            self._collision(sprite)

    def _sign_collision(self, sprite):

        if self.collide_rect.colliderect(sprite.collide_rect):

            top = Tk()

            SignWindow(top, sprite.sign_text)

            self._collision(sprite)

    def _box_collision(self, sprite):

        if self.collide_rect.colliderect(sprite.collide_rect):

            sprite.dx = self.dx

            sprite.dy = self.dy

            if self.dx > 0:
                sprite.collide_rect.left = self.collide_rect.right

            if self.dx < 0:
                sprite.collide_rect.right = self.collide_rect.left

            if self.dy > 0:
                sprite.collide_rect.top = self.collide_rect.bottom

            if self.dy < 0:
                sprite.collide_rect.bottom = self.collide_rect.top

    def _position_change_collision(self, sprite):

        if self.collide_rect.colliderect(sprite.collide_rect):

            parameters_script.BOOL = True

            current_data = sprite.get_data

            parameters_script.NEXT_MAP = os.path.join(os.getcwd(),
                                                      'ingame_data', 'maps',
                                                      current_data[0], ''.join(current_data[:]) + '.xlsx')

    def _save_point_collision(self, sprite):

        if self.collide_rect.colliderect(sprite.collide_rect):

            if PlayerDataJSON():

                t = Tk()

                t.geometry("1x1")

                if messagebox.askquestion('Внимание', 'Вы хотите перезаписать данные?\n'
                                                      'Последнее сохранение произошло:\n'
                                                      '{date}'.format(date=PlayerDataJSON().get_time)) == 'yes':

                    PlayerDataJSON().form_data(self.get_data)

                    pg.mixer.Sound.play(sprite.sound)

                t.destroy()

        self._collision(sprite)

    def _detect_collision(self):

        for wall in wall_like_sprites:

            self._collision(wall)

        for wall in gate_like_sprites:

            self._collision(wall)

        for box in movable_objects:

            self._box_collision(box)

        for chest in chests:

            self._chest_collision(chest)

        for enemy in enemies:

            self._enemy_collision(enemy)

        for sign in sign_like_sprites:

            self._sign_collision(sign)

        for position_change in change_position_things:

            self._position_change_collision(position_change)

        for save_point in save_points:

            self._save_point_collision(save_point)

    def upload_data(self, player_data):

        if player_data['TOP_LEFT_COORDINATES'] is not None:
            self.collide_rect.topleft = player_data['TOP_LEFT_COORDINATES']
        self._log_chests = player_data['LOG']
        self._numeric_stats = NumericStats(player_data['HP'], player_data['MP'])
        self._inventory = Inventory()
        for item in list(player_data['INVENTORY'].values()):
            self._inventory.append(item)

    @property
    def get_data(self):

        return parameters_thing.CURRENT_MAP,\
               self.collide_rect.topleft,\
               dt.datetime.today().strftime("%d-%m-%Y, %H:%M:%S"),\
               self._numeric_stats.get_numeric['HP'],\
               self._numeric_stats.get_numeric['MP'], \
               self._inventory.get_items, \
               self._log_chests


class EnvironmentFriendlySprite(pg.sprite.Sprite):
    """
    Потомок класса pg.sprite.Sprite.
    Необходимый спрайт для отображения текстур окружающей среды.
    """
    def __init__(self, x, y, path_to_texture, group=layered):

        pg.sprite.Sprite.__init__(self, group)

        self.image = pg.image.load(path_to_texture)

        self.rect = self.image.get_rect()

        self.rect.topleft = (x, y)


class Wall(pg.sprite.Sprite):
    """
    Потомок класса pg.sprite.Sprite.
    Является основой для остальных объектов окружающей среды, исключая EnvironmentFriendlySprite.
    Игрок не способен пройти сквозь неё.
    """
    def __init__(self, x, y, group=layered):

        pg.sprite.Sprite.__init__(self, group)

        self._layer = 1

        self.image = pg.Surface((16, 16), pg.SRCALPHA)

        self.collide_rect = pg.Rect(0, 0, TILESIZE, TILESIZE)

        self.rect = self.image.get_rect()

        self.collide_rect.topleft = (x, y)


class ChangePosition(Wall):
    """
    Уникальный объект.
    При взаимодействии игрока с данным спрайтом, загружается следующая карта,
    данные о которой хранятся в конструкторе класса.
    """
    def __init__(self, x, y):

        Wall.__init__(self, x, y)

        self.image.fill((0, 128, 255, 48))

        self.map_name = None

        self.map_id = None

        self.position = None

    def update(self):

        self.rect.bottomleft = self.collide_rect.bottomleft

    def map_data(self, data):

        self.map_name, self.map_id = data.split('|')

    @property
    def get_data(self):

        return self.map_name, self.map_id


class SavePoint(Wall):
    """
    Объект, при взаимодействии игрока с которым возможно сохранение пользовательских данных.
    """
    def __init__(self, x, y, path_to_texture):

        Wall.__init__(self, x, y)

        self.image = pg.image.load(path_to_texture)

        self.sound = pg.mixer.Sound(os.path.join(os.getcwd(), 'soundtracks', 'save.flac'))

        self.sound.set_volume(0.5)

    def update(self):

        self.rect.bottomleft = self.collide_rect.bottomleft


class Chest(Wall):
    """
    Объект типа "контейнер".
    По возможности игрок оттуда может забирать предмет, если такой имеется внутри.
    """
    def __init__(self, x, y, path_to_texture):

        Wall.__init__(self, x, y)

        self.image = pg.image.load(path_to_texture)

        self._item = None

    def update(self):

        self.rect.bottomleft = self.collide_rect.bottomleft

    def place(self, item):

        self._item = item

    def empty(self):

        self._item = None


class Gate(Wall):
    """
    Замок.
    Очень похож с объектом типа "стена", однако при взаимодействии со спрайтом "коробка" пропадает.
    """
    def __init__(self, x, y, path_to_texture):

        Wall.__init__(self, x, y)

        self.image = pg.image.load(path_to_texture)

    def update(self):

        self.rect.bottomleft = self.collide_rect.bottomleft


class Box(Wall):
    """
    Коробка ( ключ ).
    Игрок может переносить ключ по карте.
    Сам объект способен вскрывать замки.
    """
    def __init__(self, x, y, path_to_texture):

        Wall.__init__(self, x, y)

        self.dx, self.dy = 0, 0

        self.image = pg.image.load(path_to_texture)

        self.sound = pg.mixer.Sound(os.path.join(os.getcwd(), 'soundtracks', 'opened.flac'))

        self.sound.set_volume(0.5)

    def update(self):

        self.rect.bottomleft = self.collide_rect.bottomleft

        if pg.sprite.spritecollide(self, gate_like_sprites, True):

            pg.mixer.Sound.play(self.sound)

            self.kill()

        for sprite in wall_like_sprites:

            self._collision(sprite)

    def _collision(self, sprite):

        if self.collide_rect.colliderect(sprite.collide_rect):

            if self.dx > 0:
                self.collide_rect.right = sprite.collide_rect.left

            if self.dx < 0:
                self.collide_rect.left = sprite.collide_rect.right

            if self.dy > 0:
                self.collide_rect.bottom = sprite.collide_rect.top

            if self.dy < 0:
                self.collide_rect.top = sprite.collide_rect.bottom

class Sign(Wall):
    """
    Табличка.
    Хранится текст.
    И только текст.
    Примитивно.
    """
    def __init__(self, x, y, path_to_texture):

        Wall.__init__(self, x, y)

        self.image = pg.image.load(path_to_texture)

        self.sign_text = ''

    def update(self):

        self.rect.bottomleft = self.collide_rect.bottomleft

    def set_text(self, text):

        self.sign_text = text


class Enemy(Wall):
    """
    Объект типа "противник".
    Если игрок взаимодействует с подобным, начнётся сражение.
    """
    def __init__(self, x, y):

        Wall.__init__(self, x, y)

        self.enemy_type = None

        self.dx, self.dy = 0, 0

        self.velocity = np.arange(-1, 2)

    def update(self):

        self.dx, self.dy = 0, 0

        if choice([0, 1]):

            self.dx = choice(self.velocity)

            self.collide_rect.x += self.dx

        else:

            self.dy = choice(self.velocity)

            self.collide_rect.y += self.dy

        self.rect.bottomleft = self.collide_rect.bottomleft

        for group in [wall_like_sprites, gate_like_sprites, chests, sign_like_sprites]:

            for sprite in group:

                self._collision(sprite)

    def set_enemy_data(self, enemy_data):

        self.enemy_type = enemy_data

        self.image = pg.image.load(os.path.join(os.getcwd(), os.sep.join(self.enemy_type["SPRITE"])))

    def _collision(self, sprite):

        if self.collide_rect.colliderect(sprite.collide_rect):

            if self.dx > 0:
                self.collide_rect.right = sprite.collide_rect.left

            if self.dx < 0:
                self.collide_rect.left = sprite.collide_rect.right

            if self.dy > 0:
                self.collide_rect.bottom = sprite.collide_rect.top

            if self.dy < 0:
                self.collide_rect.top = sprite.collide_rect.bottom

class MapProcessor:
    """
    Класс обработки информации о карте с рабочей книги excel.
    Данные в книге хранятся следующим образом:
    Есть две таблицы с названиями "surface" и "objects".
    Первая отвечает за размещение текстур окружающей среды с помощью размещения спрайтов типа EnvironmentFriendlySprite.
    Вторая за расположение остальных объектов и проведение дополнительных опреаций над ними.
    Клетка хранит в себе следующее значение:
    INT             INT             >   INT|STR    (>   INT      (>...))
    Номер спрайта   Номер команды       Аргумент        Дополнительные аргументы
    где INT - целое число, STR - строка
    Класс поддерживает загрузку данных об игроке из JSON файла.
    """
    def __init__(self):

        self._sprites_container = [Wall, Chest, Gate, ChangePosition, Box, Sign, Enemy, SavePoint, Player]

        self._groups_container = [all_sprites_group,
                                  wall_like_sprites,
                                  chests,
                                  gate_like_sprites,
                                  change_position_things,
                                  movable_objects,
                                  sign_like_sprites,
                                  enemies,
                                  save_points,
                                  layered,
                                  player_group]

        self._field_container = [[None],
                                 ['image', 'place(self.json_item_data_processor.get_random_item_by_type({index}))'],
                                 ['image'],
                                 [None, 'map_data("{index}")'],
                                 ['image'],
                                 ['image', 'set_text(self._text_values[{index}])'],
                                 [None, 'set_enemy_data(self.json_mob_data_processor.get_mob_by_id({index}))'],
                                 ['image'],
                                 [None]]

        self._path = None

        self._name = None

        self._wb = None

        self.json_item_data_processor = ItemDataUploadJSON()

        self.json_mob_data_processor = EntityDataUploadJSON()

        self._text_values = None

        pg.mixer.init()

    def set_volume(self):

        pg.mixer.music.set_volume(parameters_thing.VOLUME)

    def _load_map(self, path):

        self._path = path

        self._name = os.path.splitext(os.path.split(self._path)[-1])[0]

        parameters_script.CURRENT_MAP = self._name

        self._text_values = open(os.path.join(os.getcwd(), 'ingame_data', 'maps', self._name[:-1], self._name + '.txt'),
                                 'r').readlines()

        self._wb = openpyxl.load_workbook(path)

        self._set_terrain()

        self._locate_objects()

        self._music_setup()

    def _load_and_check(self, map_path, player_data, flag=True):

        self._load_map(map_path)

        for sprite in chests:

            if list(sprite.collide_rect.topleft) in player_data['LOG']:

                sprite.empty()

        if not flag:

            player_data['TOP_LEFT_COORDINATES'] = None

        for sprite in all_sprites_group:

            if isinstance(sprite, Player):

                sprite.upload_data(player_data)

                break

    def json_data_load(self, player_data):

        self._load_and_check(os.path.join(os.getcwd(), 'ingame_data', 'maps', player_data['MAP'][:-1],
                                        player_data['MAP'] + '.xlsx'), player_data)

    def current_data_upload(self, player_data):

        self._load_and_check(parameters_thing.NEXT_MAP, player_data, flag=False)

    def _music_setup(self):

        parameters_script.CURRENT_SONG = os.path.join(os.getcwd(),
                                                      'ingame_data', 'maps', self._name[:-1], self._name[:-1] + '.mp3')

        load_and_play(parameters_script.CURRENT_SONG)

    def _set_terrain(self):

        surface = self._wb['surface']

        for i in range(1, surface.max_row + 1):

            for j in range(1, surface.max_column + 1):

                val = str(surface.cell(row=i, column=j).value).split()

                if all(val[i] != 'None' for i in range(len(val))):

                    for k in range(len(val)):

                        int_val = int(val[k])

                        obj = EnvironmentFriendlySprite(TILESIZE * (j - 1), TILESIZE * (i - 1),
                                                        os.path.join(os.getcwd(), 'ingame_data', 'textures',
                                                                     self._name[:-1], str(int_val) + '.png'))

                        all_sprites_group.add(obj)

    def _locate_objects(self):

        objects = self._wb['objects']

        for i in range(1, objects.max_row + 1):

            for j in range(1, objects.max_column + 1):

                val = str(objects.cell(row=i, column=j).value).split()

                if val[0] != 'None':

                    int_val = int(val[0])

                    if self._field_container[int_val - 1][0] is not None:

                        path = os.path.join(os.getcwd(), 'ingame_data', 'textures', '{}.png'.format(int_val))

                        obj = self._sprites_container[int_val - 1](TILESIZE * (j - 1), TILESIZE * (i - 1), path)

                    else:

                        obj = self._sprites_container[int_val - 1](TILESIZE * (j - 1), TILESIZE * (i - 1))

                    if len(val) > 1:

                        add_val_0 = val[1].split('>')

                        exec('obj.{}'.format(self._field_container[int_val - 1]
                                             [int(add_val_0[0])].format(index=add_val_0[1])))

                    self._groups_container[int_val].add(obj)

                    self._groups_container[0].add(obj)

    def clean_surface(self):

        for group in self._groups_container[:-1]:

            group.empty()
